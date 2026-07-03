#!/usr/bin/env python3
"""
Clinical research data de-identification (LLM-free).

Scans Excel/CSV files for Protected Health Information (PHI) using regex
and column-name heuristics, walks the researcher through an interactive
terminal review, then produces a de-identified copy with mapping and
audit trail.

Supports 10 country locales (kr, us, jp, cn, de, uk, fr, ca, au, in)
with country-specific PHI patterns.  Custom locales via --locale-file.

Usage:
    python deidentify.py scan  input.xlsx [--locale kr]
    python deidentify.py review scan_report.json
    python deidentify.py apply  reviewed_report.json [--hash-mapping]
    python deidentify.py full   input.xlsx [--locale kr] [--auto-accept-safe]
"""

import argparse
import csv
import hashlib
import json
import logging
import os
import random
import re
import stat
import sys
from datetime import datetime, timedelta
from pathlib import Path

log = logging.getLogger("deidentify")

REPORT_VERSION = 1

# ================================================================
# Section 1: Constants + Locale Loading
# ================================================================

LOCALES_DIR = Path(__file__).parent / "locales"

# Universal column names (English — common across all research locales).
UNIVERSAL_COLUMN_NAMES: dict[str, str] = {
    "patient_name": "name", "patientname": "name", "pt_name": "name",
    "name": "name", "first_name": "name", "last_name": "name",
    "ssn": "rrn", "social_security": "rrn",
    "dob": "date", "date_of_birth": "date", "birth_date": "date",
    "birthdate": "date",
    "phone": "phone", "telephone": "phone", "mobile": "phone",
    "phone_number": "phone", "cell": "phone",
    "address": "address", "home_address": "address", "street": "address",
    "zip": "address", "zipcode": "address", "zip_code": "address",
    "email": "email", "email_address": "email",
    "mrn": "id", "medical_record": "id", "chart_no": "id",
    "patient_id": "id", "patientid": "id", "chart_number": "id",
    "record_number": "id", "hospital_id": "id",
    "insurance_no": "insurance", "insurance_number": "insurance",
}

# Universal value patterns (always active regardless of locale).
UNIVERSAL_VALUE_PATTERNS: list[tuple[re.Pattern, str]] = [
    # Email
    (re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"), "email"),
    # ISO date  YYYY-MM-DD or YYYY.MM.DD or YYYY/MM/DD
    (re.compile(r"\b(19|20)\d{2}[-/.](0[1-9]|1[0-2])[-/.](0[1-9]|[12]\d|3[01])\b"), "date"),
    # YYMMDD (6 digits that look like a birthdate, standalone)
    (re.compile(r"\b([5-9]\d|0[0-4])(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])\b"), "date"),
]


def list_locales() -> list[dict]:
    """List available locales from the locales/ directory."""
    locales = []
    if not LOCALES_DIR.is_dir():
        return locales
    for f in sorted(LOCALES_DIR.glob("*.json")):
        if f.name.startswith("_"):
            continue
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            locales.append({
                "code": data.get("code", f.stem),
                "name": data.get("name", f.stem),
                "native_name": data.get("native_name", ""),
                "path": str(f),
            })
        except (json.JSONDecodeError, OSError):
            continue
    return locales


def load_locale(code: str) -> dict:
    """Load locale by country code (e.g., 'kr', 'us')."""
    path = LOCALES_DIR / f"{code}.json"
    if not path.exists():
        sys.exit(f"Locale not found: {code}\n"
                 f"Available: {', '.join(l['code'] for l in list_locales())}\n"
                 f"Or use --locale-file for a custom locale.")
    return json.loads(path.read_text(encoding="utf-8"))


def load_locale_file(path: str) -> dict:
    """Load a custom locale from an arbitrary JSON file."""
    p = Path(path)
    if not p.exists():
        sys.exit(f"Locale file not found: {path}")
    return json.loads(p.read_text(encoding="utf-8"))


def select_locale_interactive() -> dict:
    """Interactive country selection prompt."""
    locales = list_locales()
    if not locales:
        sys.exit("No locale files found in locales/ directory.")

    print(f"\n{_bold('Select country / 국가 선택:')}")
    for i, loc in enumerate(locales, 1):
        native = f" ({loc['native_name']})" if loc['native_name'] != loc['name'] else ""
        print(f"  {i:2d}. {loc['name']}{native}")
    print(f"   0. Other (provide custom locale file)")

    while True:
        choice = input(f"\n> ").strip()
        if choice == "0":
            custom_path = input("  Path to custom locale JSON: ").strip()
            locale = load_locale_file(custom_path)
            print(f"  Loaded custom locale: {locale.get('name', 'Custom')}")
            return locale
        try:
            idx = int(choice)
            if 1 <= idx <= len(locales):
                locale = load_locale(locales[idx - 1]["code"])
                print(f"  Loading {locale['name']} patterns...")
                return locale
        except ValueError:
            # Try as code
            for loc in locales:
                if choice.lower() == loc["code"]:
                    locale = load_locale(choice.lower())
                    print(f"  Loading {locale['name']} patterns...")
                    return locale
        print(f"  Invalid choice. Enter 1-{len(locales)} or a country code.")


def build_locale_patterns(locale: dict) -> tuple[
    dict[str, str],
    list[tuple[re.Pattern, str]],
    re.Pattern | None,
    re.Pattern | None,
    float,
    list[str],
]:
    """Build scanning patterns from a locale dict.

    Returns:
        (column_names, value_patterns, address_re, name_re, name_min_ratio, name_columns)
    """
    # Column names: universal + locale-specific
    column_names = dict(UNIVERSAL_COLUMN_NAMES)
    column_names.update(locale.get("column_names", {}))

    # Value patterns: universal + locale-specific
    value_patterns = list(UNIVERSAL_VALUE_PATTERNS)

    # National ID
    nid = locale.get("national_id", {})
    nid_type = nid.get("phi_type", "national_id")
    for pat in nid.get("patterns", []):
        value_patterns.append((re.compile(pat), nid_type))

    # Phone
    for phone in locale.get("phone", []):
        value_patterns.append((re.compile(phone["pattern"]), "phone"))

    # Extra date formats
    for df in locale.get("date_formats", []):
        value_patterns.append((re.compile(df["pattern"]), "date"))

    # Address pattern
    addr_cfg = locale.get("address", {})
    address_re = None
    if addr_cfg.get("type") == "suffix_regex" and addr_cfg.get("pattern"):
        address_re = re.compile(addr_cfg["pattern"])
    elif addr_cfg.get("type") == "keywords" and addr_cfg.get("keywords"):
        # Build a regex from keywords (case-insensitive word boundary match)
        escaped = [re.escape(kw) for kw in addr_cfg["keywords"]]
        address_re = re.compile(r"(?:" + "|".join(escaped) + r")", re.IGNORECASE)
    # Postcode pattern (if available, add to value_patterns as address type)
    if addr_cfg.get("postcode_pattern"):
        value_patterns.append((re.compile(addr_cfg["postcode_pattern"]), "address"))

    # Name heuristic
    name_cfg = locale.get("name_heuristic", {})
    name_re = None
    if name_cfg.get("type") == "regex" and name_cfg.get("pattern"):
        name_re = re.compile(name_cfg["pattern"])
    name_min_ratio = name_cfg.get("min_ratio", 0.3)

    # Name columns (for restricting name heuristic)
    name_columns = [k for k, v in column_names.items() if v == "name"]

    return column_names, value_patterns, address_re, name_re, name_min_ratio, name_columns

# Confidence thresholds
CONF_HIGH = "high"
CONF_MEDIUM = "medium"
CONF_LOW = "low"

# ANSI helpers (respect NO_COLOR)
_NO_COLOR = bool(os.environ.get("NO_COLOR"))


def _c(code: str, text: str) -> str:
    if _NO_COLOR:
        return text
    return f"\033[{code}m{text}\033[0m"


def _red(t: str) -> str: return _c("31", t)
def _green(t: str) -> str: return _c("32", t)
def _yellow(t: str) -> str: return _c("33", t)
def _bold(t: str) -> str: return _c("1", t)
def _dim(t: str) -> str: return _c("2", t)


# ================================================================
# Section 2: File I/O
# ================================================================

def detect_format(path: Path) -> str:
    """Return 'csv', 'tsv', or 'xlsx' based on extension."""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return "xlsx"
    if ext == ".tsv":
        return "tsv"
    if ext in (".csv", ".txt", ""):
        return "csv"
    sys.exit(f"Unsupported file format: {ext}")


def detect_encoding(path: Path) -> str:
    """Detect encoding: try UTF-8, fall back to EUC-KR."""
    raw = path.read_bytes()
    # UTF-8 BOM
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    try:
        raw.decode("euc-kr")
        return "euc-kr"
    except UnicodeDecodeError:
        pass
    return "utf-8"  # best effort


def load_tabular(path: Path) -> tuple[list[dict], dict]:
    """Load CSV/TSV/XLSX into list of row-dicts + metadata dict."""
    fmt = detect_format(path)
    meta = {"format": fmt, "path": str(path), "sheets_skipped": []}

    if fmt == "xlsx":
        try:
            import openpyxl
        except ImportError:
            sys.exit("openpyxl is required for .xlsx files.  Install: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if len(wb.sheetnames) > 1:
            meta["sheets_skipped"] = wb.sheetnames[1:]
            log.warning("Multiple sheets found. Processing '%s' only. Skipped: %s",
                        wb.sheetnames[0], ", ".join(wb.sheetnames[1:]))
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
        data = []
        for row in rows_iter:
            data.append({h: (str(v) if v is not None else "") for h, v in zip(headers, row)})
        wb.close()
    else:
        delimiter = "\t" if fmt == "tsv" else ","
        enc = detect_encoding(path)
        with open(path, newline="", encoding=enc) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            data = list(reader)

    meta["rows"] = len(data)
    meta["columns"] = len(headers) if data else 0
    meta["headers"] = headers
    return data, meta


def save_tabular(data: list[dict], path: Path, fmt: str) -> None:
    """Write de-identified data back to CSV/TSV/XLSX."""
    if not data:
        log.warning("No data to write.")
        return
    headers = list(data[0].keys())

    if fmt == "xlsx":
        try:
            import openpyxl
        except ImportError:
            fmt = "csv"
            path = path.with_suffix(".csv")
            log.warning("openpyxl not available; writing CSV instead: %s", path)

    if fmt == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append([row.get(h, "") for h in headers])
        wb.save(path)
    else:
        delimiter = "\t" if fmt == "tsv" else ","
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)


# ================================================================
# Section 3: PHI Scanner
# ================================================================

def _normalize_col(name: str) -> str:
    """Normalize column name for matching: lowercase, strip, collapse whitespace."""
    return re.sub(r"[\s_\-]+", "_", name.strip().lower())


def _col_name_matches(norm_col: str, pattern: str) -> bool:
    """Check if a normalized column name matches a PHI pattern.

    Rules:
    - Short column names (<=10 chars): exact match or pattern equals norm
    - Long column names (>10 chars): only match if pattern IS the full norm
      (prevents 'cell' matching inside 'atypical cell carcinoma')
    - Korean patterns: use substring match only for dedicated Korean PHI words
    """
    # Exact match
    if norm_col == pattern:
        return True
    # Short column name that equals or is contained in pattern
    if len(norm_col) <= 10 and norm_col in pattern:
        return True
    # Short column name: check if pattern matches as a whole word
    if len(norm_col) <= 10 and pattern in norm_col:
        return True
    # Long column name: only match if the normalized name STARTS with the pattern
    # (e.g., "전화번호_집" matches "전화번호", but "bronchial...cell" doesn't match "cell")
    if norm_col.startswith(pattern):
        return True
    return False


def scan_column_names(headers: list[str],
                      column_names: dict[str, str] | None = None) -> dict[str, dict]:
    """Match column names against PHI dictionary.

    Returns {col: {"phi_type": str, "confidence": str, "source": "column_name"}}.
    """
    if column_names is None:
        column_names = UNIVERSAL_COLUMN_NAMES
    results: dict[str, dict] = {}
    for col in headers:
        norm = _normalize_col(col)
        for pattern, phi_type in column_names.items():
            if _col_name_matches(norm, pattern):
                results[col] = {
                    "phi_type": phi_type,
                    "confidence": CONF_HIGH,
                    "source": "column_name",
                }
                break
    return results


def _sample_values(values: list[str], n: int = 500) -> list[str]:
    """Return up to n non-empty values for scanning."""
    non_empty = [v for v in values if v and v.strip()]
    if len(non_empty) <= n:
        return non_empty
    return random.sample(non_empty, n)


def scan_column_values(col: str, values: list[str],
                       col_phi_hint: str | None = None,
                       value_patterns: list[tuple[re.Pattern, str]] | None = None,
                       address_re: re.Pattern | None = None,
                       name_re: re.Pattern | None = None,
                       name_min_ratio: float = 0.3,
                       name_columns: list[str] | None = None) -> dict | None:
    """Scan cell values in a column for PHI patterns.

    Returns a detection dict or None.
    """
    if value_patterns is None:
        value_patterns = UNIVERSAL_VALUE_PATTERNS
    if name_columns is None:
        name_columns = [k for k, v in UNIVERSAL_COLUMN_NAMES.items() if v == "name"]

    sample = _sample_values(values)
    if not sample:
        return None

    # Count matches per PHI type
    type_counts: dict[str, int] = {}
    for val in sample:
        for regex, phi_type in value_patterns:
            if regex.search(val):
                type_counts[phi_type] = type_counts.get(phi_type, 0) + 1
                break  # one match per value is enough

    # Name heuristic check (only if column name hints at a name)
    if name_re is not None:
        if col_phi_hint == "name" or _normalize_col(col) in name_columns:
            name_count = sum(1 for v in sample if name_re.match(v.strip()))
            if name_count > len(sample) * name_min_ratio:
                type_counts["name"] = name_count

    # Address check
    if address_re is not None:
        addr_count = sum(1 for v in sample if address_re.search(v))
        if addr_count > len(sample) * 0.3:
            type_counts["address"] = addr_count

    if not type_counts:
        return None

    # Pick the most frequent type
    best_type = max(type_counts, key=lambda k: type_counts[k])
    ratio = type_counts[best_type] / len(sample)
    confidence = CONF_HIGH if ratio > 0.5 else CONF_MEDIUM if ratio > 0.2 else CONF_LOW

    return {
        "phi_type": best_type,
        "confidence": confidence,
        "source": "value_pattern",
        "match_ratio": round(ratio, 3),
        "sample_size": len(sample),
    }


def is_high_cardinality_numeric(values: list[str], threshold: float = 0.9) -> bool:
    """Detect columns that look like MRN/chart numbers:
    high-cardinality pure-numeric values."""
    non_empty = [v.strip() for v in values if v and v.strip()]
    if len(non_empty) < 10:
        return False
    numeric_count = sum(1 for v in non_empty if v.isdigit() and len(v) >= 5)
    if numeric_count / len(non_empty) < threshold:
        return False
    unique_ratio = len(set(non_empty)) / len(non_empty)
    return unique_ratio > 0.8


def classify_columns(data: list[dict], headers: list[str],
                      locale: dict | None = None) -> list[dict]:
    """Classify every column as PHI, SAFE, or REVIEW_NEEDED.

    Returns a list of classification dicts (one per column).
    """
    # Build patterns from locale (or use universal defaults)
    if locale is not None:
        col_names, val_patterns, addr_re, name_re, name_ratio, name_cols = \
            build_locale_patterns(locale)
    else:
        col_names = UNIVERSAL_COLUMN_NAMES
        val_patterns = UNIVERSAL_VALUE_PATTERNS
        addr_re = None
        name_re = None
        name_ratio = 0.3
        name_cols = [k for k, v in UNIVERSAL_COLUMN_NAMES.items() if v == "name"]

    # Pass 1: column name matching
    name_hits = scan_column_names(headers, col_names)

    classifications = []
    for col in headers:
        values = [row.get(col, "") for row in data]

        # Already matched by name?
        if col in name_hits:
            entry = {
                "column": col,
                "classification": "PHI",
                **name_hits[col],
            }
            # Refine with value scan
            val_hit = scan_column_values(
                col, values, name_hits[col]["phi_type"],
                val_patterns, addr_re, name_re, name_ratio, name_cols)
            if val_hit:
                entry["value_scan"] = val_hit
            classifications.append(entry)
            continue

        # Pass 2: value pattern scan
        val_hit = scan_column_values(
            col, values, None,
            val_patterns, addr_re, name_re, name_ratio, name_cols)
        if val_hit:
            classifications.append({
                "column": col,
                "classification": "PHI" if val_hit["confidence"] == CONF_HIGH else "REVIEW_NEEDED",
                **val_hit,
            })
            continue

        # Pass 3: high-cardinality numeric (possible MRN)
        if is_high_cardinality_numeric(values):
            # Show sample for user review
            unique_sample = sorted(set(v.strip() for v in values if v.strip()))[:5]
            classifications.append({
                "column": col,
                "classification": "REVIEW_NEEDED",
                "phi_type": "id",
                "confidence": CONF_LOW,
                "source": "high_cardinality_numeric",
                "sample_values": unique_sample,
            })
            continue

        # Pass 4: free-text detection (long strings, mixed content)
        non_empty = [v for v in values if v and v.strip()]
        if non_empty:
            avg_len = sum(len(v) for v in non_empty) / len(non_empty)
            if avg_len > 50:
                # Scan for embedded PHI in free text
                embedded_phi = False
                for val in _sample_values(non_empty, 100):
                    for regex, _ in val_patterns:
                        if regex.search(val):
                            embedded_phi = True
                            break
                    if embedded_phi:
                        break
                if embedded_phi:
                    classifications.append({
                        "column": col,
                        "classification": "REVIEW_NEEDED",
                        "phi_type": "free_text",
                        "confidence": CONF_MEDIUM,
                        "source": "free_text_with_phi",
                    })
                    continue

        # Default: SAFE
        classifications.append({
            "column": col,
            "classification": "SAFE",
            "phi_type": None,
            "confidence": CONF_HIGH,
            "source": "no_match",
        })

    return classifications


def build_scan_report(input_path: Path, data: list[dict],
                      meta: dict, classifications: list[dict],
                      locale: dict | None = None) -> dict:
    """Build the full scan report JSON."""
    report = {
        "version": REPORT_VERSION,
        "timestamp": datetime.now().isoformat(),
        "input_file": str(input_path),
        "meta": meta,
        "classifications": classifications,
    }
    if locale is not None:
        report["locale"] = {
            "code": locale.get("code", "custom"),
            "name": locale.get("name", "Custom"),
        }
    return report


# ================================================================
# Section 4: Interactive Reviewer
# ================================================================

def _format_classification(c: dict) -> str:
    cls = c["classification"]
    phi = c.get("phi_type", "")
    conf = c.get("confidence", "")
    if cls == "PHI":
        return f"{_red('PHI')} ({phi}, {conf})"
    if cls == "REVIEW_NEEDED":
        return f"{_yellow('REVIEW_NEEDED')} ({phi}, {conf})"
    return _green("SAFE")


def _show_sample_values(col: str, data: list[dict], n: int = 10) -> None:
    """Print up to n unique sample values for a column."""
    values = list(set(row.get(col, "") for row in data if row.get(col, "").strip()))
    sample = values[:n]
    if sample:
        print(f"  Sample values: {', '.join(repr(v) for v in sample)}")
    if len(values) > n:
        print(f"  ... and {len(values) - n} more unique values")


def review_scan_report(report: dict, data: list[dict],
                       auto_accept_safe: bool = False) -> dict:
    """Interactive three-pass review.  Mutates and returns the report."""
    classifications = report["classifications"]
    total = len(classifications)

    # ---- Pass 1: Column-level review ----
    print(f"\n{_bold('=== Pass 1: Column Classification Review ===')}")
    print(f"Total columns: {total}\n")

    phi_count = sum(1 for c in classifications if c["classification"] == "PHI")
    review_count = sum(1 for c in classifications if c["classification"] == "REVIEW_NEEDED")
    safe_count = sum(1 for c in classifications if c["classification"] == "SAFE")
    print(f"  {_red(f'PHI: {phi_count}')}  |  "
          f"{_yellow(f'REVIEW_NEEDED: {review_count}')}  |  "
          f"{_green(f'SAFE: {safe_count}')}\n")

    for i, c in enumerate(classifications):
        col = c["column"]
        cls = c["classification"]

        if cls == "SAFE" and auto_accept_safe:
            c["approved_action"] = "keep"
            continue

        print(f"[{i + 1}/{total}] {_bold(col)}: {_format_classification(c)}")
        if "sample_values" in c:
            print(f"  Flagged samples: {c['sample_values']}")
        if cls != "SAFE":
            _show_sample_values(col, data)

        if cls == "SAFE":
            choice = input("  Action [K]eep / (r)eview_needed? ").strip().lower()
            if choice == "r":
                c["classification"] = "REVIEW_NEEDED"
                c["approved_action"] = None
            else:
                c["approved_action"] = "keep"
        elif cls == "PHI":
            choice = input("  Action [A]nonymize / (k)eep / (r)eview? ").strip().lower()
            if choice == "k":
                c["approved_action"] = "keep"
            elif choice == "r":
                c["classification"] = "REVIEW_NEEDED"
                c["approved_action"] = None
            else:
                c["approved_action"] = "anonymize"
        else:  # REVIEW_NEEDED
            _show_sample_values(col, data)
            choice = input("  Action (a)nonymize / [K]eep / (f)lag_free_text? ").strip().lower()
            if choice == "a":
                c["approved_action"] = "anonymize"
                c["classification"] = "PHI"
            elif choice == "f":
                c["approved_action"] = "flag"
            else:
                c["approved_action"] = "keep"

    # ---- Pass 2: Re-examine REVIEW_NEEDED items without decisions ----
    undecided = [c for c in classifications if c.get("approved_action") is None]
    if undecided:
        print(f"\n{_bold('=== Pass 2: Undecided Items ===')}")
        for c in undecided:
            col = c["column"]
            print(f"\n  {_bold(col)}: {_format_classification(c)}")
            _show_sample_values(col, data, n=15)
            choice = input("  Action (a)nonymize / [K]eep? ").strip().lower()
            c["approved_action"] = "anonymize" if choice == "a" else "keep"

    # ---- Pass 3: Final summary ----
    print(f"\n{_bold('=== Pass 3: Final Summary ===')}")
    to_anonymize = [c for c in classifications if c.get("approved_action") == "anonymize"]
    to_keep = [c for c in classifications if c.get("approved_action") == "keep"]
    to_flag = [c for c in classifications if c.get("approved_action") == "flag"]

    print(f"\n  Anonymize ({len(to_anonymize)}): "
          + ", ".join(c["column"] for c in to_anonymize) if to_anonymize else "  Anonymize: none")
    print(f"  Keep ({len(to_keep)}): "
          + ", ".join(c["column"] for c in to_keep) if to_keep else "  Keep: none")
    if to_flag:
        print(f"  {_yellow(f'Flagged ({len(to_flag)})')}: "
              + ", ".join(c["column"] for c in to_flag))

    print()
    confirm = input("Proceed with these actions? [Y]es / (e)dit / (q)uit: ").strip().lower()
    if confirm == "q":
        sys.exit("Aborted by user.")
    if confirm == "e":
        # Allow editing individual items
        while True:
            col_name = input("  Column name to change (or 'done'): ").strip()
            if col_name.lower() == "done":
                break
            match = [c for c in classifications if c["column"] == col_name]
            if not match:
                print(f"  Column '{col_name}' not found.")
                continue
            c = match[0]
            choice = input(f"  New action for {col_name} — (a)nonymize / (k)eep / (f)lag: ").strip().lower()
            if choice == "a":
                c["approved_action"] = "anonymize"
            elif choice == "f":
                c["approved_action"] = "flag"
            else:
                c["approved_action"] = "keep"

    report["reviewed"] = True
    report["review_timestamp"] = datetime.now().isoformat()
    return report


# ================================================================
# Section 5: Anonymizers
# ================================================================

class PseudonymGenerator:
    """Maps original values to consistent pseudonyms (P001, P002, ...)."""

    def __init__(self, prefix: str = "P"):
        self._map: dict[str, str] = {}
        self._counter = 0
        self._prefix = prefix

    def get(self, original: str) -> str:
        if original not in self._map:
            self._counter += 1
            self._map[original] = f"{self._prefix}{self._counter:04d}"
        return self._map[original]

    @property
    def mapping(self) -> dict[str, str]:
        return dict(self._map)


class DateShifter:
    """Shifts dates by a consistent per-entity offset.

    The same entity (identified by entity_id) always gets the same offset,
    preserving relative time intervals between events for the same entity.
    """

    def __init__(self, seed: int, max_days: int = 365):
        self._rng = random.Random(seed)
        self._max_days = max_days
        self._offsets: dict[str, int] = {}
        self.seed = seed

    def _get_offset(self, entity_id: str) -> int:
        if entity_id not in self._offsets:
            self._offsets[entity_id] = self._rng.randint(-self._max_days, self._max_days)
        return self._offsets[entity_id]

    def shift(self, date_str: str, entity_id: str = "__default__") -> str:
        """Attempt to parse, shift, and re-format a date string."""
        offset = self._get_offset(entity_id)
        delta = timedelta(days=offset)

        # Try common formats
        for fmt_in, fmt_out in [
            ("%Y-%m-%d", "%Y-%m-%d"),
            ("%Y.%m.%d", "%Y.%m.%d"),
            ("%Y/%m/%d", "%Y/%m/%d"),
            ("%Y%m%d", "%Y%m%d"),
        ]:
            try:
                dt = datetime.strptime(date_str.strip(), fmt_in)
                return (dt + delta).strftime(fmt_out)
            except ValueError:
                continue

        # Korean format
        m = re.match(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", date_str)
        if m:
            try:
                dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                shifted = dt + delta
                return f"{shifted.year}년 {shifted.month}월 {shifted.day}일"
            except ValueError:
                pass

        # Cannot parse — return suppressed
        return "[DATE_SHIFTED]"

    @property
    def offsets(self) -> dict[str, int]:
        return dict(self._offsets)


class IDReplacer:
    """Replaces identifiers with sequential IDs (ID001, ID002, ...)."""

    def __init__(self, prefix: str = "ID"):
        self._map: dict[str, str] = {}
        self._counter = 0
        self._prefix = prefix

    def get(self, original: str) -> str:
        if original not in self._map:
            self._counter += 1
            self._map[original] = f"{self._prefix}{self._counter:04d}"
        return self._map[original]

    @property
    def mapping(self) -> dict[str, str]:
        return dict(self._map)


def _suppress(val: str) -> str:
    return "[REDACTED]"


def _sha256(val: str) -> str:
    return hashlib.sha256(val.encode("utf-8")).hexdigest()


def apply_anonymization(data: list[dict], report: dict,
                        date_shift_seed: int | None = None) -> tuple[list[dict], dict, list[dict]]:
    """Apply approved anonymization actions.

    Returns (de-identified data, mapping dict, audit entries).
    """
    classifications = report["classifications"]
    to_anonymize = {c["column"]: c for c in classifications
                    if c.get("approved_action") == "anonymize"}

    if not to_anonymize:
        log.info("No columns marked for anonymization.")
        return data, {}, []

    # Detect the entity/patient ID column for date shifting
    id_columns = [col for col, c in to_anonymize.items() if c.get("phi_type") == "id"]
    # Also check non-anonymized columns that look like IDs
    all_id_cols = id_columns + [
        c["column"] for c in classifications
        if c.get("phi_type") == "id" and c.get("approved_action") == "keep"
    ]

    # Initialize anonymizers
    name_gen = PseudonymGenerator(prefix="P")
    id_gen = IDReplacer(prefix="ID")
    seed = date_shift_seed if date_shift_seed is not None else random.randint(1, 999999)
    date_shifter = DateShifter(seed=seed)

    mapping: dict[str, dict] = {
        "_meta": {
            "date_shift_seed": seed,
            "timestamp": datetime.now().isoformat(),
            "version": REPORT_VERSION,
        }
    }
    audit: list[dict] = []

    # Process each row
    clean_data = []
    for row_idx, row in enumerate(data):
        new_row = dict(row)

        # Determine entity ID for this row (for date shifting)
        entity_id = "__default__"
        for id_col in all_id_cols:
            if row.get(id_col, "").strip():
                entity_id = row[id_col].strip()
                break

        for col, spec in to_anonymize.items():
            original = row.get(col, "")
            if not original or not original.strip():
                continue

            phi_type = spec.get("phi_type", "unknown")
            original_stripped = original.strip()

            if phi_type == "name":
                replacement = name_gen.get(original_stripped)
            elif phi_type == "id":
                replacement = id_gen.get(original_stripped)
            elif phi_type == "date":
                replacement = date_shifter.shift(original_stripped, entity_id)
            elif phi_type in ("phone", "rrn", "email", "insurance"):
                replacement = _suppress(original_stripped)
            elif phi_type == "address":
                replacement = _suppress(original_stripped)
            elif phi_type == "free_text":
                # Redact known patterns within the text
                replaced = original_stripped
                for regex, _ in PHI_VALUE_PATTERNS:
                    replaced = regex.sub("[REDACTED]", replaced)
                replacement = replaced
            else:
                replacement = _suppress(original_stripped)

            new_row[col] = replacement

            audit.append({
                "row": row_idx,
                "column": col,
                "phi_type": phi_type,
                "action": "anonymize",
                "before_hash": _sha256(original_stripped),
                "after_value": replacement,
            })

        clean_data.append(new_row)

    # Build mapping
    mapping["names"] = name_gen.mapping
    mapping["ids"] = id_gen.mapping
    mapping["date_offsets"] = date_shifter.offsets

    return clean_data, mapping, audit


# ================================================================
# Section 6: Output
# ================================================================

def write_deidentified_file(data: list[dict], input_path: Path,
                            output_dir: Path) -> Path:
    """Write de-identified data to output_dir/{stem}_deidentified.{ext}."""
    fmt = detect_format(input_path)
    out_name = f"{input_path.stem}_deidentified{input_path.suffix}"
    out_path = output_dir / out_name
    save_tabular(data, out_path, fmt)
    log.info("De-identified data written to: %s", out_path)
    return out_path


def write_mapping(mapping: dict, path: Path, hash_mode: bool = False) -> Path:
    """Write mapping file.  In hash mode, original values are SHA-256 hashed."""
    if hash_mode:
        hashed = {"_meta": mapping.get("_meta", {})}
        for section in ("names", "ids"):
            if section in mapping:
                hashed[section] = {_sha256(k): v for k, v in mapping[section].items()}
        if "date_offsets" in mapping:
            hashed["date_offsets"] = {_sha256(k): v for k, v in mapping["date_offsets"].items()}
        out_data = hashed
    else:
        out_data = mapping

    path.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")

    # Set restrictive permissions (owner-only read/write)
    try:
        os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)  # 0600
    except OSError:
        log.warning("Could not set restrictive permissions on mapping file: %s", path)

    log.info("Mapping file written to: %s (permissions: 0600)", path)
    return path


def write_audit_log(audit: list[dict], path: Path) -> Path:
    """Write audit log CSV.  before_hash is SHA-256 of original value."""
    if not audit:
        log.info("No changes made; audit log is empty.")
        return path

    fieldnames = ["row", "column", "phi_type", "action", "before_hash", "after_value"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit)

    log.info("Audit log written to: %s (%d entries)", path, len(audit))
    return path


# ================================================================
# Section 7: Main + CLI
# ================================================================

def _resolve_locale(args: argparse.Namespace) -> dict | None:
    """Resolve locale from CLI args or interactive selection."""
    if getattr(args, "locale_file", None):
        locale = load_locale_file(args.locale_file)
        log.info("Using custom locale: %s", locale.get("name", "Custom"))
        return locale
    if getattr(args, "locale", None):
        locale = load_locale(args.locale)
        log.info("Using locale: %s (%s)", locale["name"], locale["code"])
        return locale
    # Interactive selection
    return select_locale_interactive()


def cmd_scan(args: argparse.Namespace) -> None:
    """Scan command: profile and classify columns."""
    input_path = Path(args.input_file)
    if not input_path.exists():
        sys.exit(f"File not found: {input_path}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    locale = _resolve_locale(args)

    log.info("Loading %s ...", input_path)
    data, meta = load_tabular(input_path)
    log.info("Loaded %d rows, %d columns", meta["rows"], meta["columns"])

    log.info("Scanning for PHI ...")
    classifications = classify_columns(data, meta["headers"], locale)

    report = build_scan_report(input_path, data, meta, classifications, locale)
    report_path = output_dir / "scan_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    # Print summary
    phi = sum(1 for c in classifications if c["classification"] == "PHI")
    review = sum(1 for c in classifications if c["classification"] == "REVIEW_NEEDED")
    safe = sum(1 for c in classifications if c["classification"] == "SAFE")
    print(f"\n{_bold('Scan Results')}:")
    print(f"  {_red(f'PHI: {phi}')}  |  {_yellow(f'REVIEW_NEEDED: {review}')}  |  {_green(f'SAFE: {safe}')}")
    print(f"\nReport saved: {report_path}")
    print(f"Next step: python deidentify.py review {report_path}")


def cmd_review(args: argparse.Namespace) -> None:
    """Review command: interactive terminal review of scan report."""
    report_path = Path(args.report_file)
    if not report_path.exists():
        sys.exit(f"Report not found: {report_path}")

    report = json.loads(report_path.read_text(encoding="utf-8"))
    if report.get("version", 0) != REPORT_VERSION:
        log.warning("Report version mismatch (expected %d, got %d)",
                    REPORT_VERSION, report.get("version", 0))

    # Reload original data for sample display
    input_path = Path(report["input_file"])
    if not input_path.exists():
        sys.exit(f"Original file not found: {input_path}")
    data, _ = load_tabular(input_path)

    reviewed = review_scan_report(report, data,
                                  auto_accept_safe=getattr(args, "auto_accept_safe", False))

    out_path = report_path.parent / "reviewed_report.json"
    out_path.write_text(json.dumps(reviewed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nReviewed report saved: {out_path}")
    print(f"Next step: python deidentify.py apply {out_path}")


def cmd_apply(args: argparse.Namespace) -> None:
    """Apply command: anonymize based on reviewed report."""
    report_path = Path(args.report_file)
    if not report_path.exists():
        sys.exit(f"Report not found: {report_path}")

    report = json.loads(report_path.read_text(encoding="utf-8"))
    if not report.get("reviewed"):
        log.warning("Report has not been reviewed. Run 'review' first.")

    input_path = Path(report["input_file"])
    if not input_path.exists():
        sys.exit(f"Original file not found: {input_path}")

    output_dir = report_path.parent
    data, _ = load_tabular(input_path)

    log.info("Applying anonymization ...")
    clean_data, mapping, audit = apply_anonymization(data, report)

    # Write outputs
    deid_path = write_deidentified_file(clean_data, input_path, output_dir)
    mapping_path = write_mapping(mapping, output_dir / "mapping.json",
                                 hash_mode=getattr(args, "hash_mapping", False))
    audit_path = write_audit_log(audit, output_dir / "audit_log.csv")

    # Warn if mapping is in same dir as de-identified data
    if mapping_path.parent == deid_path.parent:
        print(f"\n{_yellow('WARNING')}: mapping.json is in the same directory as the "
              "de-identified data. For security, store mapping.json separately.")

    # Summary
    changes = len(audit)
    cols_changed = len(set(a["column"] for a in audit))
    print(f"\n{_bold('De-identification Complete')}:")
    print(f"  Changes: {changes} cells across {cols_changed} columns")
    print(f"  Output:  {deid_path}")
    print(f"  Mapping: {mapping_path}")
    print(f"  Audit:   {audit_path}")


def cmd_full(args: argparse.Namespace) -> None:
    """Full pipeline: scan -> review -> apply in one go."""
    input_path = Path(args.input_file)
    if not input_path.exists():
        sys.exit(f"File not found: {input_path}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    locale = _resolve_locale(args)

    # Scan
    log.info("Loading %s ...", input_path)
    data, meta = load_tabular(input_path)
    log.info("Loaded %d rows, %d columns", meta["rows"], meta["columns"])

    log.info("Scanning for PHI ...")
    classifications = classify_columns(data, meta["headers"], locale)
    report = build_scan_report(input_path, data, meta, classifications, locale)

    # Quick summary before review
    phi = sum(1 for c in classifications if c["classification"] == "PHI")
    review_n = sum(1 for c in classifications if c["classification"] == "REVIEW_NEEDED")
    safe = sum(1 for c in classifications if c["classification"] == "SAFE")
    print(f"\n{_bold('Scan Results')}:")
    print(f"  {_red(f'PHI: {phi}')}  |  {_yellow(f'REVIEW_NEEDED: {review_n}')}  |  {_green(f'SAFE: {safe}')}")

    if phi == 0 and review_n == 0:
        print(f"\n{_green('No PHI detected.')} Your data appears clean.")
        confirm = input("Proceed anyway? (y/n) ").strip().lower()
        if confirm != "y":
            return

    # Review
    reviewed = review_scan_report(report, data,
                                  auto_accept_safe=args.auto_accept_safe)

    # Save report
    report_path = output_dir / "reviewed_report.json"
    report_path.write_text(json.dumps(reviewed, ensure_ascii=False, indent=2), encoding="utf-8")

    # Apply
    log.info("Applying anonymization ...")
    clean_data, mapping, audit = apply_anonymization(data, reviewed)

    # Write outputs
    deid_path = write_deidentified_file(clean_data, input_path, output_dir)
    mapping_path = write_mapping(mapping, output_dir / "mapping.json",
                                 hash_mode=args.hash_mapping)
    audit_path = write_audit_log(audit, output_dir / "audit_log.csv")

    if mapping_path.parent == deid_path.parent:
        print(f"\n{_yellow('WARNING')}: mapping.json is in the same directory as the "
              "de-identified data. For security, store mapping.json separately.")

    changes = len(audit)
    cols_changed = len(set(a["column"] for a in audit))
    print(f"\n{_bold('De-identification Complete')}:")
    print(f"  Changes: {changes} cells across {cols_changed} columns")
    print(f"  Output:  {deid_path}")
    print(f"  Mapping: {mapping_path}")
    print(f"  Audit:   {audit_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="deidentify",
        description="Clinical research data de-identification (LLM-free).",
    )
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Enable verbose logging")
    sub = parser.add_subparsers(dest="command", required=True)

    # Locale options (shared by scan and full)
    def _add_locale_args(p: argparse.ArgumentParser) -> None:
        g = p.add_mutually_exclusive_group()
        g.add_argument("--locale", type=str, metavar="CODE",
                       help="Country code (kr, us, jp, cn, de, uk, fr, ca, au, in). "
                            "If omitted, interactive selection is shown.")
        g.add_argument("--locale-file", type=str, metavar="PATH",
                       help="Path to a custom locale JSON file")

    # scan
    p_scan = sub.add_parser("scan", help="Scan a file for PHI")
    p_scan.add_argument("input_file", help="Path to CSV/TSV/XLSX file")
    p_scan.add_argument("-o", "--output-dir", default=".", help="Output directory (default: .)")
    _add_locale_args(p_scan)

    # review
    p_review = sub.add_parser("review", help="Interactive review of scan report")
    p_review.add_argument("report_file", help="Path to scan_report.json")
    p_review.add_argument("--auto-accept-safe", action="store_true",
                          help="Automatically accept SAFE columns without prompting")

    # apply
    p_apply = sub.add_parser("apply", help="Apply anonymization from reviewed report")
    p_apply.add_argument("report_file", help="Path to reviewed_report.json")
    p_apply.add_argument("--hash-mapping", action="store_true",
                         help="Hash original values in mapping file (one-way)")

    # full
    p_full = sub.add_parser("full", help="Full pipeline: scan + review + apply")
    p_full.add_argument("input_file", help="Path to CSV/TSV/XLSX file")
    p_full.add_argument("-o", "--output-dir", default=".", help="Output directory (default: .)")
    _add_locale_args(p_full)
    p_full.add_argument("--auto-accept-safe", action="store_true",
                        help="Automatically accept SAFE columns without prompting")
    p_full.add_argument("--hash-mapping", action="store_true",
                        help="Hash original values in mapping file (one-way)")

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    if args.command == "scan":
        cmd_scan(args)
    elif args.command == "review":
        cmd_review(args)
    elif args.command == "apply":
        cmd_apply(args)
    elif args.command == "full":
        cmd_full(args)


if __name__ == "__main__":
    main()
